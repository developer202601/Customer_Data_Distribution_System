#!/usr/bin/env python3
import sys
import json
from openpyxl import load_workbook

def main():
    if len(sys.argv) < 2:
        print(json.dumps({
            'status': 'error',
            'message': 'Usage: read_payment_excel.py <file.xlsx>'
        }))
        sys.exit(1)

    file_path = sys.argv[1]

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(json.dumps({
            'status': 'error',
            'message': f'Failed to open workbook: {e}'
        }))
        sys.exit(1)

    sheet = wb.active
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    rows = []
    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        if any(v is not None and str(v).strip() != '' for v in values):
            rows.append(values)

    wb.close()

    if not rows:
        print(json.dumps({
            'status': 'error',
            'message': 'No data found in workbook.'
        }))
        sys.exit(1)

    header_row = [str(h).strip() if h is not None else '' for h in rows[0]]
    upper_headers = [h.upper() for h in header_row]

    account_num_index = None
    payment_mny_index = None

    for idx, header in enumerate(upper_headers):
        if header == 'ACCOUNT_NUM':
            account_num_index = idx
        if header == 'ACCOUNT_PAYMENT_MNY':
            payment_mny_index = idx

    if account_num_index is None or payment_mny_index is None:
        found = ', '.join(header_row[:20])
        print(json.dumps({
            'status': 'error',
            'message': f"Missing required headers. Found: [{found}]"
        }))
        sys.exit(1)

    data_rows = []
    for row in rows[1:]:
        data_rows.append(list(row))

    print(json.dumps({
        'status': 'ok',
        'headers': header_row,
        'account_num_index': account_num_index,
        'payment_mny_index': payment_mny_index,
        'total_rows': len(data_rows),
        'data': data_rows
    }))

if __name__ == '__main__':
    main()
